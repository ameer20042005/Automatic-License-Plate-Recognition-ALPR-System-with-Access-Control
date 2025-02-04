import os
import cv2
import serial
import time
import threading
from openpyxl import load_workbook
from fast_alpr import ALPR

# تحقق من وجود ملف Excel
excel_path = "car_numbers.xlsx"
if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    exit()

# قراءة ملف Excel
wb = load_workbook(excel_path)
ws = wb.active

# إعداد الاتصال مع Arduino (تأكد من المنفذ الصحيح مثل COM3 أو /dev/ttyUSB0)
arduino = serial.Serial('COM6', 9600, timeout=1)
time.sleep(2)  # انتظار استقرار الاتصال

# تهيئة ALPR باستخدام وحدة المعالجة المركزية فقط
alpr = ALPR(
    detector_model="yolo-v9-t-384-license-plate-end2end",
    ocr_model="global-plates-mobile-vit-v2-model",
)

# فتح الكاميرا
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("Error: Could not open camera.")
    exit()

# دالة للتحقق إذا كان رقم اللوحة موجودًا في ملف Excel
def check_plate_in_excel(plate_number):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == plate_number:
            return True
    return False

# دالة لفتح وإغلاق السيرفو في Thread منفصل
def control_servo():
    print("Opening Servo at 100 degrees...")
    arduino.write(b'100\n')  # إرسال زاوية 100 إلى Arduino
    time.sleep(6)  # الانتظار لمدة 6 ثوانٍ دون تجميد البرنامج
    print("Closing Servo (return to 0 degrees)...")
    arduino.write(b'0\n')  # إرجاع المحرك إلى زاوية 0

# قراءة الإطارات ومعالجتها
try:
    previous_plate_number = None  # لحفظ الرقم السابق الذي تم التحقق منه

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error: Could not read frame.")
            break

        # الكشف عن اللوحات
        alpr_results = alpr.predict(frame)
        detected_plate_number = None

        for result in alpr_results:
            if hasattr(result, "ocr") and result.ocr is not None:
                detected_plate_number = result.ocr.text.strip()
                break  # استخراج أول رقم لوحة فقط

        if detected_plate_number:
            print(f"Detected Plate Number: {detected_plate_number}")

            # تحقق من الرقم في ملف Excel
            if detected_plate_number != previous_plate_number:
                previous_plate_number = detected_plate_number  # تحديث الرقم السابق
                if check_plate_in_excel(detected_plate_number):
                    print("Plate number found in Excel.")
                    threading.Thread(target=control_servo).start()  # تشغيل السيرفو بدون تعطيل الكاميرا
                else:
                    print("Plate number not found in Excel.")

        # رسم التنبؤات على الإطار
        annotated_frame = alpr.draw_predictions(frame)

        # عرض النتائج
        cv2.imshow("ALPR Camera", annotated_frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # إيقاف الفيديو عند الضغط على 'q'
            break
except Exception as e:
    print(f"Error during processing: {e}")
finally:
    # تحرير الموارد
    cap.release()
    cv2.destroyAllWindows()
    arduino.close()
    print("تم إيقاف تشغيل الكاميرا وإغلاق الاتصال بـ Arduino.")
